import io
import base64
import re
import time as _time
import hmac
import hashlib
import html as _html
from datetime import date as _date
import streamlit as st
import extra_streamlit_components as stx
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from PIL import Image as PILImage

st.set_page_config(
    page_title="Wholesale Quotation · Chhajed Garden",
    page_icon="🌿",
    layout="wide",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&display=swap');

/* ── Full width, no constraints ── */
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; font-size: 16px; }
.main .block-container {
    padding: 2rem 3rem 4rem 3rem !important;
    max-width: 1100px !important;
    width: 100% !important;
}
h1, h2, h3 { font-family: 'DM Serif Display', serif; }

/* ── Header ── */
.app-header {
    background: linear-gradient(135deg, #1f5c2e 0%, #2d7a40 100%);
    color: white; padding: 2rem 2.5rem; border-radius: 16px;
    margin-bottom: 2rem; display: flex; align-items: center; gap: 1.5rem;
}
.app-header h1 { color: white; margin: 0; font-size: 2.2rem; }
.app-header p  { color: #c8e6c9; margin: 0.3rem 0 0; font-size: 1.1rem; }

/* ── Order card ── */
.order-card {
    background: white; border: 1px solid #e0e0e0; border-radius: 12px;
    padding: 1.8rem; margin-bottom: 1.5rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}
.order-card h3 { margin-top: 0; color: #1f5c2e; font-size: 1.3rem; }
.info-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 0.6rem 2rem; font-size: 1.1rem; }
.info-row { display: flex; gap: 0.5rem; }
.info-label { color: #888; min-width: 140px; font-size: 1.05rem; }
.info-value { color: #222; font-weight: 500; font-size: 1.1rem; }

/* ── Section title ── */
.section-title {
    font-family: 'DM Serif Display', serif;
    color: #1f5c2e; font-size: 1.4rem; font-weight: 600;
    margin: 1.8rem 0 1rem;
}

/* ── All selectbox dropdowns ── */
[data-testid="stSelectbox"] label,
[data-testid="stNumberInput"] label {
    font-size: 1.1rem !important; font-weight: 500 !important; color: #333 !important;
}
/* Target every possible container inside Streamlit's selectbox */
[data-testid="stSelectbox"] > div:last-child { overflow: visible !important; }
[data-testid="stSelectbox"] > div:last-child > div { overflow: visible !important; }
[data-baseweb="select"] {
    min-height: 56px !important; overflow: visible !important;
}
[data-baseweb="select"] > div {
    min-height: 56px !important; overflow: visible !important;
    padding: 12px 14px !important; align-items: center !important;
    display: flex !important; flex-wrap: nowrap !important;
}
[data-baseweb="select"] > div > div {
    min-height: 56px !important; overflow: visible !important;
    display: flex !important; align-items: center !important;
}
[data-baseweb="select"] * {
    font-size: 1.15rem !important; line-height: 1.6 !important;
}

/* ── Number input ── */
[data-testid="stNumberInput"] input {
    font-size: 1.15rem !important; min-height: 56px !important; padding: 12px 14px !important;
}

/* ── Totals ── */
.totals-block {
    background: white; border: 1px solid #e0e0e0;
    border-radius: 12px; overflow: hidden; margin-top: 1rem;
}
.totals-row {
    display: flex; justify-content: space-between;
    padding: 14px 24px; border-bottom: 1px solid #f0f0f0; font-size: 1.1rem;
}
.totals-row:last-child { border-bottom: none; }
.totals-row.final { background: #1f5c2e; color: white; font-weight: 700; font-size: 1.2rem; }
.totals-label { font-weight: 500; }
.totals-value { font-weight: 600; }

/* ── Uploader ── */
[data-testid="stFileUploader"] {
    border: 2px dashed #1f5c2e44; border-radius: 12px;
    padding: 1rem; background: #f0f7f1;
}

/* ── Download button ── */
[data-testid="stDownloadButton"] button {
    background: #1f5c2e !important; color: white !important;
    border-radius: 8px !important; font-weight: 600 !important;
    padding: 0.7rem 2rem !important; font-size: 1.1rem !important; border: none !important;
}
[data-testid="stDownloadButton"] button:hover { background: #174d25 !important; }
.stSpinner > div { border-top-color: #1f5c2e !important; }
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
GREEN      = "1F5C2E"   # section banners, headers
LGREEN     = "E8F5E9"   # subtotal rows
WHITE      = "FFFFFF"
GREY       = "F5F5F5"
# Option B · Tonal Greens
PALE_SAGE  = "D8EDD8"   # logistics rows (destination, transport, billing)
VQUOTE_BG  = "EEF5EE"   # V1–V7 quote rows (very pale sage)
KYELLOW    = "FFFF00"   # booking type row

IMG_ROW_HEIGHT = 175
IMG_COL_WIDTH  = 30.5

TAXABLE_KEYWORDS = ["sculpture", "ceramic", "spray pump", "cocopeat", "moss", "media"]

HEADER_FIELDS = [
    ("Order ID",          "Order ID"),
    ("Customer Name",     "Customer Name"),
    ("Customer Phone",    "Customer Phone"),
    ("City",              "City"),
    ("State",             "State"),
    ("Country",           "Country"),
    ("Pincode",           "Pincode"),
    ("Complete address",  "Complete address"),
]

TRANSPORT_MODES = ["Train", "Travels", "Air"]
PACKING_TYPES   = ["Newspaper wrap (₹3/plant)", "Newspaper + cardboard box (₹5/plant)"]
STAFF_NAMES     = ["Riya", "Sanika", "Shreya", "Preeti", "Reshma", "Sonal"]
QUOTE_VERSIONS  = [f"V{i}" for i in range(1, 8)]


def ordinal_date(d) -> str:
    """Return e.g. '14th May 2026'."""
    n = d.day
    suffix = (
        "th" if 11 <= n <= 13 else
        {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    )
    return f"{n}{suffix} {d.strftime('%B %Y')}"


def short_date(d) -> str:
    """Return e.g. '14-May-26' for Excel cells."""
    return d.strftime("%-d-%b-%y")


def is_taxable(name: str) -> bool:
    n = name.lower()
    return any(kw in n for kw in TAXABLE_KEYWORDS)


def is_ceramic_pot(name: str) -> bool:
    n = name.lower()
    if re.search(r'\d+\s*cm\s+pot', n):  # plant sold in a pot — not a standalone pot
        return False
    return bool(re.search(r'\bceramic\b|\bpots?\b', n))


def get_set_size(name: str) -> int:
    m = re.search(r'\(S/(\d+)\)|S/(\d+)', name, re.I)
    if m:
        return int(m.group(1) or m.group(2))
    return 1


# ── Parser (reused from retail app) ───────────────────────────────────────────

def parse_xlsx(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    kv = {}
    product_header_row = None
    col_idx = {}

    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]

        if "Product Name" in cells and ("Product SKU" in cells or "Quantity" in cells):
            product_header_row = i
            for field in ("Product Name", "Product SKU", "Product Price",
                          "Discounted Price", "Price", "MRP", "Selling Price", "Quantity"):
                if field in cells:
                    col_idx[field] = cells.index(field)
            break

        non_empty = [c for c in row if c is not None]
        if len(non_empty) >= 2:
            kv[str(non_empty[0]).strip()] = non_empty[1]
        elif len(non_empty) == 1:
            kv[str(non_empty[0]).strip()] = None

    if product_header_row is None:
        raise ValueError("Could not find product table (looking for 'Product Name' header row).")

    name_col  = col_idx.get("Product Name")
    sku_col   = col_idx.get("Product SKU")
    price_col = (col_idx.get("Product Price") or col_idx.get("Discounted Price")
                 or col_idx.get("Price") or col_idx.get("MRP") or col_idx.get("Selling Price"))
    qty_col   = col_idx.get("Quantity")

    if name_col is None:
        raise ValueError("Could not find 'Product Name' column in product table.")

    img_by_row = {}
    for img in ws._images:
        try:
            row_idx = img.anchor._from.row
            img.ref.seek(0)
            data = img.ref.read()
            img_by_row[row_idx] = (data, img.format or "jpeg")
        except Exception:
            pass

    products = []
    for abs_row, row in enumerate(rows[product_header_row + 1:], product_header_row + 1):
        def cell(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        name  = cell(name_col)
        sku   = cell(sku_col)
        price = cell(price_col)
        qty   = cell(qty_col)

        if name is None and sku is None:
            break

        name  = str(name).strip() if name else ""
        sku   = str(sku).strip()  if sku  else ""

        try:
            price = float(str(price).replace(",", "").strip()) if price not in (None, "", "None") else 0
        except (ValueError, TypeError):
            price = 0

        try:
            qty = int(float(str(qty).strip())) if qty not in (None, "", "None") else 1
        except (ValueError, TypeError):
            qty = 1

        img_data = img_by_row.get(abs_row)
        products.append({"name": name, "sku": sku, "price": price, "qty": qty, "img": img_data})

    return {"kv": kv, "products": products}


def parse_pdf(file_bytes: bytes) -> dict:
    import pdfplumber

    all_lines = []
    product_images = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            all_lines.extend((page.extract_text() or "").splitlines())
            page_imgs = sorted(
                [img for img in page.images if img["x0"] > 50],
                key=lambda x: x["top"],
            )
            for img in page_imgs:
                try:
                    product_images.append((img["stream"].get_data(), "jpeg"))
                except Exception:
                    product_images.append(None)

    full_text = "\n".join(all_lines)
    kv = {}
    m = re.search(r'Order ID[:\s]+(\S+)', full_text)
    if m:
        kv["Order ID"] = m.group(1)

    cust_idx = next((i for i, l in enumerate(all_lines) if "CUSTOMER DETAILS" in l), None)
    prod_hdr_idx = next((i for i, l in enumerate(all_lines) if "No. Product Item" in l), None)
    if cust_idx is not None and prod_hdr_idx is not None:
        section = all_lines[cust_idx + 1 : prod_hdr_idx]
        noise_re = re.compile(r'(TOTAL|Estimate|₹|products|\d{4}$|^India$)', re.I)
        phone_re = re.compile(r'^\+\d')
        name_candidates = [l.strip() for l in section
                           if l.strip() and not phone_re.match(l.strip()) and not noise_re.search(l.strip())]
        kv["Customer Name"] = name_candidates[0] if name_candidates else ""

        phone = next((l.strip() for l in section if phone_re.match(l.strip())), "")
        kv["Customer Phone"] = phone

        city_state_re = re.compile(r'^(.+?),\s*(.+?)\s*-\s*(\d{6})')
        for l in section:
            cm = city_state_re.match(l.strip())
            if cm:
                kv["City"]    = cm.group(1).strip()
                kv["State"]   = cm.group(2).strip()
                kv["Pincode"] = cm.group(3).strip()
                break

        addr_lines = []
        skip_name = kv.get("Customer Name", "")
        for l in section:
            l = l.strip()
            if not l or l == skip_name or phone_re.match(l) or l in ("India", "Estimate"):
                continue
            if re.match(r'^Order ID', l, re.I):
                continue
            l = re.sub(r'\s*TOTAL\s*$', '', l)
            l = re.sub(r'\s*₹\s*[\d,]+', '', l)
            l = re.sub(r'\d+\s*products', '', l).strip().rstrip(',')
            if l:
                addr_lines.append(l)
        kv["Complete address"] = ", ".join(addr_lines)

    product_re = re.compile(r'^(\d+)\s+(.+?)\s+(\d+)\s+₹\s*([\d,]+)\s+₹\s*([\d,]+)\s*$')
    sku_re     = re.compile(r'^SKU\s*:\s*(\S+)')

    products = []
    for i, line in enumerate(all_lines):
        pm = product_re.match(line.strip())
        if pm:
            _, name, qty, price, _ = pm.groups()
            sku = ""
            if i + 1 < len(all_lines):
                sm = sku_re.match(all_lines[i + 1].strip())
                if sm:
                    sku = sm.group(1)
            idx = len(products)
            products.append({
                "name":  name.strip(),
                "sku":   sku,
                "price": float(price.replace(",", "")),
                "qty":   int(qty),
                "img":   product_images[idx] if idx < len(product_images) else None,
            })

    if not products:
        raise ValueError("No products found in PDF. Is this a Quicksell estimate PDF?")

    return {"kv": kv, "products": products}


def img_to_b64(data: bytes, fmt: str) -> str:
    mime = "image/jpeg" if fmt.lower() in ("jpg", "jpeg") else f"image/{fmt.lower()}"
    return f"data:{mime};base64,{base64.b64encode(data).decode()}"


# ── Packing cost calculator ────────────────────────────────────────────────────

def compute_packing_cost(packing_cfg: dict, products: list) -> float:
    ptype     = packing_cfg["type"]
    total_qty = sum(p["qty"] for p in products)

    # Ceramic pots: ₹10 × set_size × qty per pot product
    ceramic_products = [p for p in products if is_ceramic_pot(p["name"])]
    if ceramic_products:
        cost = 0
        for p in ceramic_products:
            cost += 10 * get_set_size(p["name"]) * p["qty"]
        # Non-ceramic products in same order use newspaper wrap by default
        non_ceramic_qty = sum(p["qty"] for p in products if not is_ceramic_pot(p["name"]))
        if ptype == "Newspaper + cardboard box (₹5/plant)":
            cost += non_ceramic_qty * 5
        else:
            cost += non_ceramic_qty * 3
        return cost

    if ptype == "Newspaper wrap (₹3/plant)":
        return total_qty * 3
    elif ptype == "Newspaper + cardboard box (₹5/plant)":
        return total_qty * 5

    return 0


def compute_transport_cost(transport_cfg: dict, products: list) -> float:
    mode      = transport_cfg["mode"]
    total_qty = sum(p["qty"] for p in products)
    if mode in ("Train", "Travels"):
        return total_qty * 15
    elif mode == "Air":
        return 1000
    return 0


# ── Excel builder ──────────────────────────────────────────────────────────────

def build_xlsx(data: dict, plant_packing: float, plant_transport: float,
               plant_packing_label: str, pot_packing: float,
               transport_mode: str, taxable_map: dict = None,
               quote_version: str = "V1", quote_by: str = "",
               quote_date: str = "") -> bytes:
    kv       = data["kv"]
    products = data["products"]
    if taxable_map is None:
        taxable_map = {p["name"]: is_taxable(p["name"]) for p in products}

    thin = Side(style="thin", color="1F5C2E")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fl(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    wb = Workbook()
    ws = wb.active
    ws.title = "Wholesale Quotation"

    col_widths = [6, IMG_COL_WIDTH, 42, 18, 13, 13, 10, 13]
    for col, w in zip(range(1, 9), col_widths):
        ws.column_dimensions[get_column_letter(col)].width = w

    def hdr_cell(r, col, text, bold=False, size=10, color="000000",
                 ha="center", va="center", wrap=False, fill=None):
        c = ws.cell(r, col, text)
        c.font      = Font(name="Arial", bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
        if fill:
            c.fill = fl(fill)
        return c

    def sub_banner(label):
        nonlocal row
        for col in range(1, 9):
            ws.cell(row, col).fill = fl(GREEN)
        c = ws.cell(row, 2, label)
        c.font      = Font(name="Arial", bold=True, size=8, color=WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 14
        row += 1

    def sage_row(label, val="", val2=""):
        nonlocal row
        for col in range(1, 9):
            ws.cell(row, col).fill = fl(PALE_SAGE)
        hdr_cell(row, 2, label, bold=True, fill=PALE_SAGE)
        if val:
            hdr_cell(row, 3, val, fill=PALE_SAGE)
        if val2:
            hdr_cell(row, 4, val2, bold=True, fill=PALE_SAGE)
        ws.row_dimensions[row].height = 20
        row += 1
        return row - 1  # return row number just written

    row = 1

    # ── Customer info rows (white) ─────────────────────────────────────────────
    info_fields = [
        ("Customer Name",  kv.get("Customer Name", "")),
        ("Customer Phone", kv.get("Customer Phone", "")),
        ("City",           kv.get("City", "")),
        ("State",          kv.get("State", "")),
        ("Country",        kv.get("Country", "India")),
    ]
    for label, val in info_fields:
        hdr_cell(row, 2, label, bold=True)
        hdr_cell(row, 3, val)
        ws.row_dimensions[row].height = 18
        row += 1

    # Complete address (tall row, white)
    hdr_cell(row, 2, "Complete address", bold=True, va="top")
    hdr_cell(row, 3, kv.get("Complete address", ""), va="top", wrap=True)
    ws.row_dimensions[row].height = 90
    row += 1

    # ── LOGISTICS & BILLING section ───────────────────────────────────────────
    sub_banner("LOGISTICS & BILLING")

    # Destination row
    sage_row("Destination", kv.get("City", ""))

    # Mode of Transport row (pre-filled, dropdown in cell C)
    transport_display = "By Air" if transport_mode == "Air" else transport_mode
    transport_row_num = sage_row("Mode of Transport", transport_display)

    # Billing Name row
    sage_row("Billing Name", kv.get("Customer Name", ""), "Sanjay Nursery")

    # ── QUOTES section ─────────────────────────────────────────────────────────
    sub_banner("QUOTES")

    # V1–V7 Quote rows (pale sage)
    selected_v = int(quote_version[1]) if quote_version else 1
    for v in range(1, 8):
        for col in range(1, 9):
            ws.cell(row, col).fill = fl(VQUOTE_BG)
        hdr_cell(row, 2, f"V{v} Quote", bold=True, fill=VQUOTE_BG)
        if v == selected_v:
            hdr_cell(row, 3, quote_by,   fill=VQUOTE_BG)
            hdr_cell(row, 4, quote_date, fill=VQUOTE_BG)
        ws.row_dimensions[row].height = 20
        row += 1

    # Booking Type row (sage green, dropdown)
    for col in range(1, 9):
        ws.cell(row, col).fill = fl(PALE_SAGE)
    hdr_cell(row, 2, "Booking Type", bold=True, fill=PALE_SAGE)
    c_booking = ws.cell(row, 3, "To Pay / Paid")
    c_booking.font      = Font(name="Arial", size=10, bold=False)
    c_booking.alignment = Alignment(horizontal="center", vertical="center")
    c_booking.fill      = fl(PALE_SAGE)
    booking_row_num = row
    ws.row_dimensions[row].height = 20
    row += 2  # blank spacer before product table

    # ── Data Validation dropdowns ─────────────────────────────────────────────
    dv_transport = DataValidation(
        type="list",
        formula1='"Train,Travels,Air,Truck"',
        allow_blank=True, showDropDown=False,
    )
    dv_transport.sqref = f"C{transport_row_num}"
    ws.add_data_validation(dv_transport)

    dv_booking = DataValidation(
        type="list",
        formula1='"To Pay / Paid,To Pay,Paid,Advance Paid"',
        allow_blank=True, showDropDown=False,
    )
    dv_booking.sqref = f"C{booking_row_num}"
    ws.add_data_validation(dv_booking)

    # ── Helper: write one column-header row ──────────────────────────────────
    def write_col_headers(plain=False):
        nonlocal row
        headers = ["Sr No", "Product Image", "Product Name", "Product SKU",
                   "Retail Price", "Wholesale Price", "Quantity", "Amount"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row, col, h)
            if plain:
                c.font = Font(name="Arial", bold=True, color="000000", size=10)
                c.fill = fl(WHITE)
            else:
                c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
                c.fill = fl(GREEN)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = bdr
        ws.row_dimensions[row].height = 22
        row += 1

    def section_banner(title, plain=False):
        nonlocal row
        for col in range(1, 9):
            ws.cell(row, col).fill = fl(WHITE if plain else GREEN)
        c = ws.cell(row, 1, title)
        c.font      = Font(name="Arial", bold=True,
                           color="000000" if plain else WHITE, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 26
        row += 1

    def write_products(product_list, plain=False):
        nonlocal row
        start = row
        for i, p in enumerate(product_list, 1):
            taxable  = taxable_map.get(p["name"], is_taxable(p["name"]))
            row_fill = fl(WHITE) if plain else (fl(GREY) if i % 2 == 0 else fl(WHITE))
            ws.row_dimensions[row].height = IMG_ROW_HEIGHT
            retail_price    = p["price"]
            wholesale_price = round(retail_price / 2, 2)
            vals   = [i, "", p["name"], p["sku"], retail_price, wholesale_price,
                      p["qty"], f"=F{row}*G{row}"]
            aligns = ["center", "center", "left", "left",
                      "center", "center", "center", "center"]
            for col, (val, ha) in enumerate(zip(vals, aligns), 1):
                c = ws.cell(row, col, val)
                c.font      = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=(col == 3))
                c.fill      = row_fill
                c.border    = bdr
            if taxable and not plain:
                ws.cell(row, 3).fill = fl("FFF8E1")
            if p["img"]:
                try:
                    img_data, fmt = p["img"]
                    pil = PILImage.open(io.BytesIO(img_data)).convert("RGB")
                    pil.thumbnail((200, 200), PILImage.LANCZOS)
                    buf = io.BytesIO()
                    pil.save(buf, format="JPEG", quality=95)
                    buf.seek(0)
                    xl_img = XLImage(buf)
                    EMU_PER_PX = 9525
                    pad = 8 * EMU_PER_PX
                    r0 = row - 1
                    ci = 1
                    _from  = AnchorMarker(col=ci,     colOff=pad,  row=r0,     rowOff=pad)
                    _to    = AnchorMarker(col=ci + 1, colOff=-pad, row=r0 + 1, rowOff=-pad)
                    anchor = TwoCellAnchor(editAs="twoCell")
                    anchor._from = _from
                    anchor.to    = _to
                    xl_img.anchor = anchor
                    ws.add_image(xl_img)
                except Exception:
                    pass
            row += 1
        return start, row - 1

    def total_row(label, qty_val, amt_val, bold=False, bg=None):
        nonlocal row
        for col in range(1, 9):
            c = ws.cell(row, col)
            c.border = bdr
            c.font   = Font(name="Arial", bold=bold, size=10,
                            color=WHITE if bg == GREEN else "000000")
            if bg:
                c.fill = fl(bg)
        ws.cell(row, 3, label).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 3).font = Font(name="Arial", bold=bold, size=10,
                                    color=WHITE if bg == GREEN else "000000")
        if qty_val is not None:
            ws.cell(row, 7, qty_val).alignment = Alignment(horizontal="center", vertical="center")
        if amt_val is not None:
            ws.cell(row, 8, amt_val).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

    # Split by taxable status (mirrors the UI split)
    plant_products_xl = [p for p in products if not taxable_map.get(p["name"], is_taxable(p["name"]))]
    other_products_xl = [p for p in products if taxable_map.get(p["name"], is_taxable(p["name"]))]
    has_both_xl       = bool(plant_products_xl) and bool(other_products_xl)

    plant_subtotal_xl = sum(round(p["price"] / 2, 2) * p["qty"] for p in plant_products_xl)
    other_subtotal_xl = sum(round(p["price"] / 2, 2) * p["qty"] for p in other_products_xl)
    plant_total_xl    = plant_subtotal_xl + plant_packing + plant_transport
    other_total_xl    = other_subtotal_xl + pot_packing

    plant_qty_xl = sum(p["qty"] for p in plant_products_xl)

    # ── Plants section (plain white style) ───────────────────────────────────
    if plant_products_xl:
        section_banner("PLANTS QUOTE", plain=True)
        write_col_headers(plain=True)
        p_start, p_end = write_products(plant_products_xl, plain=True)

        total_row("Subtotal", f"=SUM(G{p_start}:G{p_end})",
                  f"=SUM(H{p_start}:H{p_end})", bold=True)

        if transport_mode in ("Train", "Travels"):
            pt_formula = f"={plant_qty_xl}*15"
        else:
            rate       = 3 if "₹3" in plant_packing_label else 5
            pt_formula = f"={plant_qty_xl}*{rate}+1000"
        total_row("Packing & Transport", None, pt_formula, bold=True)

        total_row("Plants Total" if has_both_xl else "Final Amount",
                  None, plant_total_xl, bold=True)

    # ── Pots & Accessories section (plain white style) ────────────────────────
    if other_products_xl:
        row += 1
        section_banner("POTS & ACCESSORIES QUOTE", plain=True)
        write_col_headers(plain=True)
        o_start, o_end = write_products(other_products_xl, plain=True)

        total_row("Subtotal", f"=SUM(G{o_start}:G{o_end})",
                  f"=SUM(H{o_start}:H{o_end})", bold=True)

        ceramic_lines_xl = []
        for p in other_products_xl:
            if is_ceramic_pot(p["name"]):
                set_sz = get_set_size(p["name"])
                cost   = 10 * set_sz * p["qty"]
                ceramic_lines_xl.append(
                    f"  {p['name']}: ₹10 × {set_sz}(set) × {p['qty']} = ₹{cost:,}"
                    if set_sz > 1 else
                    f"  {p['name']}: ₹10 × {p['qty']} = ₹{cost:,}")
        total_row("Packing", None, pot_packing, bold=True)

        total_row("Pots Total" if has_both_xl else "Final Amount",
                  None, other_total_xl, bold=True)

    # ── Grand total ────────────────────────────────────────────────────────────
    if has_both_xl:
        row += 1
        total_row("Grand Total", None, plant_total_xl + other_total_xl, bold=True)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ── Auth ───────────────────────────────────────────────────────────────────────

SESSION_TIMEOUT_MINS = 30
COOKIE_NAME          = "cg_auth"
_cm                  = stx.CookieManager(key="cg_cookie_mgr")

def _app_password() -> str:
    pw = st.secrets.get("APP_PASSWORD", "")
    if not pw:
        st.error("APP_PASSWORD secret is not configured. Set it in Streamlit secrets.")
        st.stop()
    return pw

def _sign(timestamp: str, secret: str) -> str:
    return hmac.new(secret.encode(), timestamp.encode(), hashlib.sha256).hexdigest()

def _make_token(secret: str) -> str:
    ts = str(int(_time.time()))
    return f"{ts}.{_sign(ts, secret)}"

def _verify_token(token: str, secret: str) -> bool:
    try:
        ts, sig = token.rsplit(".", 1)
        if not hmac.compare_digest(_sign(ts, secret), sig):
            return False
        return _time.time() - int(ts) <= SESSION_TIMEOUT_MINS * 60
    except Exception:
        return False

def check_password() -> bool:
    cm     = _cm
    secret = _app_password()
    token  = cm.get(COOKIE_NAME)

    if token:
        if _verify_token(token, secret):
            st.session_state["authenticated"] = True
            return True
        cm.delete(COOKIE_NAME)
        st.session_state["authenticated"] = False
        st.warning("Session expired. Please log in again.")

    if st.session_state.get("authenticated"):
        return True

    st.markdown("""
    <div style="max-width:400px; margin: 8rem auto; text-align:center;">
      <div style="font-size:3rem; margin-bottom:1rem;">🌿</div>
      <h2 style="font-family:'DM Serif Display',serif; color:#1f5c2e; margin-bottom:0.25rem;">
        Wholesale Quotation
      </h2>
      <p style="color:#888; margin-bottom:2rem; font-size:0.95rem;">
        Chhajed Garden · Sanjay Nursery
      </p>
    </div>
    """, unsafe_allow_html=True)

    col = st.columns([1, 2, 1])[1]
    with col:
        pwd = st.text_input("Password", type="password", placeholder="Enter password…", label_visibility="collapsed")
        if pwd:
            if hmac.compare_digest(pwd, secret):
                cm.set(COOKIE_NAME, _make_token(secret), max_age=SESSION_TIMEOUT_MINS * 60)
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    return False

if not check_password():
    st.stop()


# ── UI ─────────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="app-header">
  <div style="font-size:2.8rem">🌿</div>
  <div>
    <h1>Wholesale Quotation Generator</h1>
    <p>Chhajed Garden · Sanjay Nursery — Upload a Quicksell export to generate a wholesale quotation</p>
  </div>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Upload Quicksell order export (.xlsx or .pdf)",
    type=["xlsx", "pdf"],
    label_visibility="collapsed",
    help="Upload a Quicksell .xlsx export or .pdf estimate"
)

if uploaded:
    with st.spinner("Reading order..."):
        try:
            file_bytes = uploaded.read()
            if uploaded.name.lower().endswith(".pdf"):
                data = parse_pdf(file_bytes)
            else:
                data = parse_xlsx(file_bytes)
        except Exception:
            st.error("Could not read file. Please check it is a valid Quicksell export.")
            st.stop()

    kv = data["kv"]

    # Assign stable unique _id based on original position (handles duplicate SKUs)
    for idx, p in enumerate(data["products"]):
        p["_id"] = idx

    # Initialise tax status in session_state on first load
    for p in data["products"]:
        sk = f"tax_{p['_id']}"
        if sk not in st.session_state:
            st.session_state[sk] = "GST" if is_taxable(p["name"]) else "No Tax"

    # Sort using current session_state values so changes re-order immediately
    def _taxable_now(p):
        return st.session_state.get(f"tax_{p['_id']}", "No Tax") == "GST"

    products         = sorted(data["products"], key=_taxable_now)
    data["products"] = products
    total_qty        = sum(p["qty"] for p in products)

    # ── Order info card ────────────────────────────────────────────────────────
    st.markdown(f"""
    <div class="order-card">
      <h3>📦 Order {kv.get('Order ID','—')}</h3>
      <div class="info-grid">
        <div class="info-row"><span class="info-label">Customer</span><span class="info-value">{kv.get('Customer Name','—')}</span></div>
        <div class="info-row"><span class="info-label">Phone</span><span class="info-value">{kv.get('Customer Phone','—')}</span></div>
        <div class="info-row"><span class="info-label">City</span><span class="info-value">{kv.get('City','—')}, {kv.get('State','—')}</span></div>
        <div class="info-row"><span class="info-label">Pincode</span><span class="info-value">{kv.get('Pincode','—')}</span></div>
        <div class="info-row" style="grid-column:1/-1"><span class="info-label">Address</span><span class="info-value">{kv.get('Complete address','—')}</span></div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Transport & Packing configuration ────────────────────────────────────
    st.markdown('<p class="section-title">Transport & Packing</p>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        transport_mode = st.selectbox("Mode of Transport", TRANSPORT_MODES)
        if transport_mode == "Air":
            st.info("Air: ₹5/plant packing + ₹1,000 flat to airport")
        else:
            st.info("Train / Travels: ₹15/plant covers packing + transport")

    with col2:
        if transport_mode in ("Train", "Travels"):
            packing_type  = "train_allin"
            packing_label = "Included in transport (₹15/plant all-in)"
        else:
            packing_type  = st.selectbox("Packing Type", PACKING_TYPES)
            packing_label = packing_type

    col3, col4 = st.columns(2)
    with col3:
        quote_version = st.selectbox("Quote Version", QUOTE_VERSIONS)
    with col4:
        quote_by = st.selectbox("Quote By", STAFF_NAMES)

    today         = _date.today()
    date_long     = ordinal_date(today)    # "14th May 2026"
    date_short_xl = short_date(today)      # "14-May-26"

    # ── Split: non-taxable = plants, taxable = pots/accessories ──────────────
    plant_products = [p for p in products if not _taxable_now(p)]
    pot_products   = [p for p in products if _taxable_now(p)]

    plant_qty      = sum(p["qty"] for p in plant_products)
    plant_subtotal = sum(round(p["price"] / 2, 2) * p["qty"] for p in plant_products)
    pot_subtotal   = sum(round(p["price"] / 2, 2) * p["qty"] for p in pot_products)

    if transport_mode in ("Train", "Travels"):
        plant_transport     = plant_qty * 15
        plant_packing       = 0
        plant_packing_label = "Included in transport (₹15/plant all-in)"
    else:  # Air
        plant_transport     = 1000 if plant_products else 0
        plant_packing       = plant_qty * 3 if packing_type == "Newspaper wrap (₹3/plant)" else plant_qty * 5
        plant_packing_label = packing_type

    plant_total = plant_subtotal + plant_packing + plant_transport
    # Only ceramic pots within the taxable group attract ₹10/pot packing
    pot_packing = sum(10 * get_set_size(p["name"]) * p["qty"]
                      for p in pot_products if is_ceramic_pot(p["name"]))
    pot_total   = pot_subtotal + pot_packing
    grand_total = plant_total + pot_total

    # ── Product table ─────────────────────────────────────────────────────────
    TABLE_HDR = """
    <div style="display:grid;grid-template-columns:0.4fr 1.2fr 4fr 1.8fr 1.8fr 1.1fr 1.2fr 0.8fr 1.2fr;
                background:#1f5c2e;color:white;padding:14px 16px;
                font-size:1rem;font-weight:600;letter-spacing:0.05em;text-transform:uppercase;gap:8px;">
      <div>#</div><div>IMAGE</div><div>PRODUCT NAME</div><div>TAX</div><div>SKU</div>
      <div>RETAIL</div><div>WHOLESALE</div><div>QTY</div><div>AMOUNT</div>
    </div>"""

    IMG_PX = 168
    MT     = "margin-top:62px"

    taxable_map = {}

    def render_product_rows(product_list, start_num=1):
        for i, p in enumerate(product_list, start_num):
            wholesale_price = round(p["price"] / 2, 2)
            amount = wholesale_price * p["qty"]

            st.markdown("<hr style='margin:0;border:none;border-top:1px solid #e8e8e8'>",
                        unsafe_allow_html=True)
            c_num, c_img, c_name, c_tax, c_sku, c_retail, c_ws, c_qty, c_amt = \
                st.columns([0.4, 1.2, 4.0, 1.8, 1.8, 1.1, 1.2, 0.8, 1.2])

            with c_num:
                st.markdown(f"<p style='color:#bbb;font-size:1.1rem;text-align:center;"
                            f"margin-top:{IMG_PX//2 - 10}px'>{i}</p>", unsafe_allow_html=True)
            with c_img:
                if p["img"]:
                    try:
                        img_data, _ = p["img"]
                        pil = PILImage.open(io.BytesIO(img_data)).convert("RGB")
                        pil.thumbnail((IMG_PX, IMG_PX), PILImage.LANCZOS)
                        st.image(pil, width=IMG_PX)
                    except Exception:
                        st.markdown(f"<div style='width:{IMG_PX}px;height:{IMG_PX}px;"
                                    f"background:#f0f0f0;border-radius:8px'></div>",
                                    unsafe_allow_html=True)
                else:
                    st.markdown(f"<div style='width:{IMG_PX}px;height:{IMG_PX}px;background:#f0f0f0;"
                                f"border-radius:8px;display:flex;align-items:center;"
                                f"justify-content:center;font-size:2.5rem'>🌱</div>",
                                unsafe_allow_html=True)
            with c_name:
                st.markdown(f"<p style='font-size:1.25rem;font-weight:500;line-height:1.5;{MT}'>"
                            f"{_html.escape(p['name'])}</p>", unsafe_allow_html=True)
            with c_tax:
                st.markdown(f"<p style='{MT}'></p>", unsafe_allow_html=True)
                sk      = f"tax_{p['_id']}"
                current = st.session_state.get(sk, "GST" if is_taxable(p["name"]) else "No Tax")
                tax = st.selectbox("tax", ["No Tax", "GST"],
                                   index=1 if current == "GST" else 0,
                                   key=sk, label_visibility="collapsed")
                taxable_map[p["name"]] = (tax == "GST")
            with c_sku:
                st.markdown(f"<p style='font-family:monospace;color:#555;font-size:1.1rem;{MT}'>"
                            f"{_html.escape(p['sku'])}</p>", unsafe_allow_html=True)
            with c_retail:
                st.markdown(f"<p style='text-align:center;font-size:1.2rem;{MT}'>"
                            f"₹{p['price']:,.0f}</p>", unsafe_allow_html=True)
            with c_ws:
                st.markdown(f"<p style='text-align:center;font-size:1.2rem;font-weight:600;{MT}'>"
                            f"₹{wholesale_price:,.0f}</p>", unsafe_allow_html=True)
            with c_qty:
                st.markdown(f"<p style='text-align:center;font-size:1.2rem;{MT}'>"
                            f"{p['qty']}</p>", unsafe_allow_html=True)
            with c_amt:
                st.markdown(f"<p style='text-align:center;font-weight:700;color:#1f5c2e;"
                            f"font-size:1.2rem;{MT}'>₹{amount:,.0f}</p>", unsafe_allow_html=True)

    # Plants section
    if plant_products:
        st.markdown('<p class="section-title">Plants</p>', unsafe_allow_html=True)
        st.markdown(TABLE_HDR, unsafe_allow_html=True)
        render_product_rows(plant_products, start_num=1)

    # Pots & Accessories section
    if pot_products:
        st.markdown('<p class="section-title" style="margin-top:2rem">Pots & Accessories</p>',
                    unsafe_allow_html=True)
        st.markdown(TABLE_HDR, unsafe_allow_html=True)
        render_product_rows(pot_products, start_num=1)

    # ── Ceramic packing breakdown ─────────────────────────────────────────────
    ceramic_lines = []
    for p in pot_products:
        if is_ceramic_pot(p["name"]):
            set_sz = get_set_size(p["name"])
            cost   = 10 * set_sz * p["qty"]
            if set_sz > 1:
                ceramic_lines.append(
                    f"{p['name']} — ₹10 × {set_sz} (set) × {p['qty']} = ₹{cost:,}")
            else:
                ceramic_lines.append(
                    f"{p['name']} — ₹10 × {p['qty']} = ₹{cost:,}")

    # ── Totals ────────────────────────────────────────────────────────────────
    has_both = bool(plant_products) and bool(pot_products)

    if plant_products:
        pt_combined = plant_packing + plant_transport
        if transport_mode in ("Train", "Travels"):
            pt_detail = f"{transport_mode} all-in: ₹15 × {plant_qty} plants = ₹{pt_combined:,.0f}"
        else:
            rate = 3 if "₹3" in plant_packing_label else 5
            pt_detail = (f"Packing ₹{rate}/plant × {plant_qty} = ₹{plant_packing:,.0f}"
                         f" + Air ₹{plant_transport:,.0f} = ₹{pt_combined:,.0f}")
        st.markdown(f"""
        <div class="totals-block" style="margin-top:1.5rem">
          <div style="padding:12px 24px;background:#f0f7f1;font-weight:600;font-size:1.1rem;
                      color:#1f5c2e;border-bottom:1px solid #e0e0e0;">
            Plants Quote ({len(plant_products)} items)
          </div>
          <div class="totals-row">
            <span class="totals-label">Subtotal</span>
            <span class="totals-value">₹{plant_subtotal:,.0f}</span>
          </div>
          <div class="totals-row" style="align-items:flex-start">
            <span class="totals-label">Packing &amp; Transport
              <br><small style="color:#888;font-size:0.9rem">{pt_detail}</small>
            </span>
            <span class="totals-value" style="padding-top:2px">₹{pt_combined:,.0f}</span>
          </div>
          <div class="totals-row final">
            <span class="totals-label">{'Plants Total' if has_both else 'Final Amount'}</span>
            <span class="totals-value">₹{plant_total:,.0f}</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

    if pot_products:
        st.markdown(f"""
        <div class="totals-block" style="margin-top:1.5rem">
          <div style="padding:12px 24px;background:#f0f7f1;font-weight:600;font-size:1.1rem;
                      color:#1f5c2e;border-bottom:1px solid #e0e0e0;">
            Pots Quote ({len(pot_products)} items)
          </div>
          <div class="totals-row">
            <span class="totals-label">Wholesale Subtotal</span>
            <span class="totals-value">₹{pot_subtotal:,.0f}</span>
          </div>
          <div class="totals-row" style="align-items:flex-start">
            <span class="totals-label">
              Packing — ₹10/pot
              {''.join(f'<br><small style="color:#888;font-size:0.9rem">{line}</small>' for line in ceramic_lines)}
            </span>
            <span class="totals-value" style="padding-top:2px">₹{pot_packing:,.0f}</span>
          </div>
          <div class="totals-row final">
            <span class="totals-label">{'Pots Total' if has_both else 'Final Amount'}</span>
            <span class="totals-value">₹{pot_total:,.0f}</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

    if has_both:
        st.markdown(f"""
        <div class="totals-block" style="margin-top:1.5rem">
          <div class="totals-row final" style="border-radius:12px">
            <span class="totals-label">Grand Total</span>
            <span class="totals-value">₹{grand_total:,.0f}</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    with st.spinner("Building Excel file..."):
        xlsx_bytes = build_xlsx(data, plant_packing, plant_transport, plant_packing_label,
                                pot_packing, transport_mode, taxable_map,
                                quote_version=quote_version, quote_by=quote_by,
                                quote_date=date_short_xl)

    customer_name = kv.get('Customer Name', 'Wholesale')
    file_name = f"{quote_version} Quote for {customer_name} dated {date_long} by {quote_by}.xlsx"
    st.download_button(
        label="⬇️  Download Wholesale Quotation (.xlsx)",
        data=xlsx_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.markdown("""
    <div style="text-align:center; padding: 3rem 1rem; color: #999;">
      <div style="font-size:3rem; margin-bottom:1rem">📂</div>
      <p style="font-size:1rem;">Upload a Quicksell .xlsx or .pdf export above to get started</p>
    </div>
    """, unsafe_allow_html=True)
